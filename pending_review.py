# pending_review.py

import os
import atexit
from openpyxl import load_workbook, Workbook
from config import PENDING_REVIEW_FILE, INPUT_EXCEL_PATH, PENDING_FLUSH_INTERVAL

# In-memory buffer of pending-review records
_records = []

def get_city_sheet_name():
    """Derive the city sheet name from the input Excel file name."""
    base = os.path.basename(INPUT_EXCEL_PATH)
    name, _ = os.path.splitext(base)
    parts = name.split()
    if parts and parts[0].upper() == 'FINAL':
        city = ' '.join(parts[1:])
    else:
        city = name
    return city[:31]

def _flush_records(records):
    """Write the provided records to the Excel sheet."""
    sheet_name = get_city_sheet_name()
    if os.path.exists(PENDING_REVIEW_FILE):
        wb = load_workbook(PENDING_REVIEW_FILE)
    else:
        wb = Workbook()
        default = wb.active
        if default and default.title == 'Sheet':
            wb.remove(default)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(['ActivityId','Name','Type','Website','Google_Map_Link','PhotoCount'])
    for rec in records:
        ws.append([
            rec['ActivityId'],
            rec['Name'],
            rec['Type'],
            rec['Website'],
            rec['Google_Map_Link'],
            rec['PhotoCount']
        ])
    wb.save(PENDING_REVIEW_FILE)

def add_pending_review(record):
    """Buffer a pending-review record, flushing periodically."""
    _records.append(record)
    if len(_records) >= PENDING_FLUSH_INTERVAL:
        _flush_records(_records)
        _records.clear()

def flush_pending_reviews():
    """Flush any remaining buffered records to the Excel sheet."""
    if _records:
        _flush_records(_records)
        _records.clear()

# Ensure flush on program exit
atexit.register(flush_pending_reviews)
